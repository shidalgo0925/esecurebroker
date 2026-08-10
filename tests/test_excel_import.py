"""Assisted Excel import + pilot seed — synthetic fixtures only (no real PII)."""

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from corredores.db import Base, SessionLocal, engine
from corredores.domain import models as _models  # noqa: F401
from corredores.domain.enums import DataSource
from corredores.domain.models import CommissionRule, Installment, Policy
from corredores.services.excel_import import (
    EmissionImportRow,
    PartyImportRow,
    run_assisted_import,
)
from corredores.services.excel_xlsx import load_emisiones_xlsx
from corredores.services.installment_status import derive_installment_status
from corredores.services.seed_pilot import seed_pilot


def setup_module():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def test_seed_pilot_commission_rates():
    with SessionLocal() as session:
        report = seed_pilot(session)
        session.commit()
        rules = session.query(CommissionRule).all()
    assert report["commission_rules_created"] >= 1
    assert any(r.rate == Decimal("0.20") for r in rules)
    assert any(r.agreement_reference == "PILOTO_EXCEL_FORMULA_V1" for r in rules)


def test_import_auto_emitido_and_coexistence():
    parties = [
        PartyImportRow(
            row_number=2,
            first_name="Demo",
            last_name="Cliente",
            national_id="8-999-1",
        )
    ]
    emissions = [
        EmissionImportRow(
            row_number=2,
            excel_status="EMITIDO",
            carrier_name="SURA",
            risk_label="AUTOMÓVIL",
            first_name="Demo",
            last_name="Cliente",
            national_id="8-999-1",
            policy_number="POL-SYN-001",
            make="TOYOTA",
            model="HILUX",
            year=2024,
            plate="SYN001",
            usage="PARTICULAR",
            premium=Decimal("100.00"),
            annual_premium=Decimal("1200.00"),
            num_payments=12,
            installment_amounts=[Decimal("100.00")] + [None] * 11,
            installment_paid_hints=[True] + [None] * 11,
            effective_date=date(2026, 1, 1),
            expiration_date=date(2027, 1, 1),
        ),
        EmissionImportRow(
            row_number=3,
            excel_status="EMITIDO",
            carrier_name="FEDPA",
            risk_label="VIDA INDIVIDUAL",
            first_name="Demo",
            last_name="Cliente",
            national_id="8-999-1",
            policy_number="VIDA-SYN-1",
            premium=Decimal("50.00"),
            annual_premium=Decimal("50.00"),
            num_payments=1,
            effective_date=date(2026, 2, 1),
        ),
        EmissionImportRow(
            row_number=4,
            excel_status="GESTIONADO",
            carrier_name="SURA",
            risk_label="AUTOMÓVIL",
            first_name="Demo",
            last_name="Cliente",
            national_id="8-999-1",
        ),
    ]
    with SessionLocal() as session:
        r1 = run_assisted_import(session, parties=parties, emissions=emissions)
        session.commit()
        # Coexistence: re-import same policy number must skip.
        r2 = run_assisted_import(session, emissions=[emissions[0]])
        session.commit()
        policies = session.query(Policy).all()
        auto = [p for p in policies if p.policy_number == "POL-SYN-001"][0]
        assert auto.data_source == DataSource.EXCEL_IMPORT
        inst = session.query(Installment).filter_by(payment_plan_id=auto.payment_plan.id).first()
        st = derive_installment_status(inst, today=date(2026, 8, 10))

    assert r1.policies_created == 2
    assert r1.payments_from_hint == 1
    assert r1.tasks_created == 1
    assert r2.policies_skipped_existing == 1
    assert st.value == "PAID"


def test_red_hint_does_not_create_overdue_status():
    """D-02: red/False hint must not invent OVERDUE — unpaid remains PENDING/DUE by date."""
    emissions = [
        EmissionImportRow(
            row_number=10,
            excel_status="EMITIDO",
            carrier_name="ANCON",
            risk_label="AUTO",
            first_name="X",
            last_name="Y",
            national_id="1-1-1",
            policy_number="POL-RED-1",
            premium=Decimal("100"),
            annual_premium=Decimal("100"),
            num_payments=1,
            installment_amounts=[Decimal("100")],
            installment_paid_hints=[False],  # red ignored
            effective_date=date(2026, 9, 1),  # future → PENDING
        )
    ]
    with SessionLocal() as session:
        run_assisted_import(session, emissions=emissions)
        session.commit()
        pol = session.query(Policy).filter_by(policy_number="POL-RED-1").one()
        inst = session.query(Installment).filter_by(payment_plan_id=pol.payment_plan.id).one()
        st = derive_installment_status(inst, today=date(2026, 8, 10))
    assert st.value == "PENDING"


def test_xlsx_loader_synthetic(tmp_path: Path):
    path = tmp_path / "emisiones_syn.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"
    ws.append(
        [
            "Estatus",
            "Cia",
            "Riesgo",
            "Nombre",
            "Apellido",
            "Cédula/RUC",
            "No Póliza",
            "Marca",
            "Modelo",
            "Año",
            "Placa",
            "Uso Auto",
            "Prima",
            "Prima Anual",
            "No. Pagos",
            "1",
            "Vig. Inicial",
            "Vig. Final",
            "Pago",
        ]
    )
    ws.append(
        [
            "EMITIDO",
            "SURA",
            "AUTOMÓVIL",
            "Syn",
            "User",
            "9-9-9",
            "XLSX-1",
            "KIA",
            "RIO",
            2022,
            "ABC123",
            "PARTICULAR",
            80,
            960,
            12,
            80,
            date(2026, 1, 1),
            date(2027, 1, 1),
            "NORMAL",
        ]
    )
    # Green fill on cuota 1
    from openpyxl.styles import PatternFill

    ws.cell(row=2, column=16).fill = PatternFill("solid", fgColor="00B050")
    wb.save(path)

    rows = load_emisiones_xlsx(path)
    assert len(rows) == 1
    assert rows[0].policy_number == "XLSX-1"
    assert rows[0].installment_paid_hints[0] is True
