"""XLSX adapters for assisted import — optional openpyxl.

Never log cell values that look like PII; callers get typed rows only.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from corredores.services.excel_import import EmissionImportRow, PartyImportRow, _as_date, _dec


def _cell(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _header_index(headers: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = str(h or "").strip()
        if not key:
            continue
        # First occurrence wins for duplicate headers (Cia / Referido / Tipo).
        out.setdefault(key, i)
        out.setdefault(key.casefold(), i)
    return out


def _idx(hmap: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in hmap:
            return hmap[name]
        if name.casefold() in hmap:
            return hmap[name.casefold()]
    return None


def _year(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value)))
    except ValueError:
        return None


def _fill_color_is_green(cell: Any) -> bool | None:
    """Return True only for known paid-green; never treat red as overdue."""
    try:
        fill = cell.fill
        color = getattr(getattr(fill, "fgColor", None), "rgb", None) or getattr(
            getattr(fill, "start_color", None), "rgb", None
        )
        if not color:
            return None
        c = str(color).upper()
        if c.endswith("00B050") or c in {"FF00B050", "0000B050"}:
            return True
        # Explicit ignore of red — caller must not invent OVERDUE.
        if c.endswith("FF0000") or c in {"FFFF0000", "00FF0000"}:
            return False
    except Exception:
        return None
    return None


def load_asegurados_xlsx(path: Path | str) -> list[PartyImportRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c or "").strip() for c in next(rows_iter)]
    h = _header_index(headers)
    out: list[PartyImportRow] = []
    for n, raw in enumerate(rows_iter, start=2):
        if not any(raw):
            continue
        out.append(
            PartyImportRow(
                row_number=n,
                first_name=str(_cell(raw, _idx(h, "Nombre")) or "").strip() or None,
                last_name=str(_cell(raw, _idx(h, "Apellido")) or "").strip() or None,
                national_id=str(_cell(raw, _idx(h, "Cédula / Ruc", "Cédula/Ruc", "Cedula")) or "").strip()
                or None,
                phone=str(_cell(raw, _idx(h, "Teléfono", "Telefono")) or "").strip() or None,
                district=str(_cell(raw, _idx(h, "Distrito")) or "").strip() or None,
                address=str(_cell(raw, _idx(h, "Dirección", "Direccion")) or "").strip() or None,
                birth_date=_as_date(_cell(raw, _idx(h, "F. Nacimiento", "F Nacimiento"))),
            )
        )
    return out


def load_emisiones_xlsx(path: Path | str) -> list[EmissionImportRow]:
    from openpyxl import load_workbook

    # data_only for values; second pass without data_only for fill colors on cuota cols.
    wb_vals = load_workbook(path, data_only=True)
    ws_vals = wb_vals["Registro"] if "Registro" in wb_vals.sheetnames else wb_vals[wb_vals.sheetnames[0]]
    wb_fmt = load_workbook(path, data_only=False)
    ws_fmt = wb_fmt["Registro"] if "Registro" in wb_fmt.sheetnames else wb_fmt[wb_fmt.sheetnames[0]]

    header_row = next(ws_vals.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(c or "").strip() for c in header_row]
    h = _header_index(headers)

    # Cuota columns often named "1".."12"
    cuota_idxs = [(_idx(h, str(i)), i) for i in range(1, 13)]
    cuota_idxs = [(idx, n) for idx, n in cuota_idxs if idx is not None]

    out: list[EmissionImportRow] = []
    for n, raw in enumerate(ws_vals.iter_rows(min_row=2, values_only=True), start=2):
        if not any(raw):
            continue
        amounts: list[Decimal | None] = [None] * 12
        hints: list[bool | None] = [None] * 12
        fmt_row = list(ws_fmt.iter_rows(min_row=n, max_row=n))[0]
        for idx, num in cuota_idxs:
            amounts[num - 1] = _dec(_cell(raw, idx))
            hints[num - 1] = _fill_color_is_green(fmt_row[idx])

        year_val = _year(_cell(raw, _idx(h, "Año", "Ano")))
        out.append(
            EmissionImportRow(
                row_number=n,
                excel_status=str(_cell(raw, _idx(h, "Estatus")) or "").strip(),
                carrier_name=str(_cell(raw, _idx(h, "Cia")) or "").strip() or None,
                risk_label=str(_cell(raw, _idx(h, "Riesgo")) or "").strip() or None,
                coverage_type=str(_cell(raw, _idx(h, "Tipo Cobertura")) or "").strip() or None,
                contractor_name=str(
                    _cell(raw, _idx(h, "Compañía/Contratante", "Compania/Contratante")) or ""
                ).strip()
                or None,
                first_name=str(_cell(raw, _idx(h, "Nombre")) or "").strip() or None,
                last_name=str(_cell(raw, _idx(h, "Apellido")) or "").strip() or None,
                national_id=str(
                    _cell(raw, _idx(h, "Cédula/RUC", "Cédula / RUC", "Cedula/RUC")) or ""
                ).strip()
                or None,
                policy_number=str(_cell(raw, _idx(h, "No Póliza", "No Poliza")) or "").strip() or None,
                make=str(_cell(raw, _idx(h, "Marca")) or "").strip() or None,
                model=str(_cell(raw, _idx(h, "Modelo")) or "").strip() or None,
                year=year_val,
                plate=str(_cell(raw, _idx(h, "Placa")) or "").strip() or None,
                vehicle_type=str(_cell(raw, _idx(h, "Tipo")) or "").strip() or None,
                usage=str(_cell(raw, _idx(h, "Uso Auto")) or "").strip() or None,
                premium=_dec(_cell(raw, _idx(h, "Prima"))),
                annual_premium=_dec(_cell(raw, _idx(h, "Prima Anual"))),
                num_payments=_year(_cell(raw, _idx(h, "No. Pagos", "No Pagos"))),
                installment_amounts=amounts,
                installment_paid_hints=hints,
                payment_form=str(_cell(raw, _idx(h, "Forma Pago")) or "").strip() or None,
                pago_column=str(_cell(raw, _idx(h, "Pago")) or "").strip() or None,
                effective_date=_as_date(_cell(raw, _idx(h, "Vig. Inicial", "Vig Inicial"))),
                expiration_date=_as_date(_cell(raw, _idx(h, "Vig. Final", "Vig Final"))),
                registro_date=_as_date(_cell(raw, _idx(h, "Registro", "Fecha"))),
            )
        )
    return out


def load_workbook_bundle(
    *,
    asegurados: Path | str | None = None,
    emisiones: Path | str | None = None,
) -> tuple[list[PartyImportRow], list[EmissionImportRow]]:
    parties: list[PartyImportRow] = []
    emissions: list[EmissionImportRow] = []
    if asegurados:
        parties = load_asegurados_xlsx(asegurados)
    if emisiones:
        emissions = load_emisiones_xlsx(emisiones)
    return parties, emissions
